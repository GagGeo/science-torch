"""
core/excel_manager.py — Gestion du fichier Excel modulaire
Crée et met à jour le fichier Excel avec onglets thématiques dynamiques.
"""
import sys as _sys
import os as _os
_sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))


import re
from datetime import datetime
from pathlib import Path
from typing import Optional
from itertools import combinations

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from utils.logger import get_logger

logger = get_logger(__name__)

# ── Palette de couleurs ───────────────────────────────────────────────────────
COLORS = {
    "header_bg":      "2C3E50",  # Bleu-gris foncé
    "header_fg":      "FFFFFF",
    "subheader_bg":   "3498DB",  # Bleu
    "row_even":       "EBF5FB",
    "row_odd":        "FFFFFF",
    "tab_experimental":"2ECC71",  # Vert
    "tab_review":     "E67E22",  # Orange
    "tab_combined":   "9B59B6",  # Violet
    "tab_all":        "1ABC9C",  # Turquoise
    "unread":         "FADBD8",  # Rose pâle
    "read":           "D5F5E3",  # Vert pâle
    "pdf_ok":         "A9DFBF",
    "pdf_missing":    "F9EBEA",
}

COLUMNS_EXPERIMENTAL = [
    ("Référence",                   40),
    ("Année",                        8),
    ("Clé citable",                 15),
    ("BibTeX",                      30),
    ("Type",                        20),
    ("Hypothèse(s)",                40),
    ("Population",                  25),
    ("N par groupe",                15),
    ("Type de groupe",              25),
    ("Critères incl./excl.",        35),
    ("Méthode / Outils",            35),
    ("Résultats principaux",        45),
    ("Taille d'effet",              15),
    ("Conclusion",                  40),
    ("Take Home Message",           40),
    ("Domaines",                    20),
    ("Lu",                          8),
    ("PDF disponible",              15),
    ("Chemin PDF",                  40),
    ("Date ajout",                  15),
    ("PMID",                        12),
    ("DOI",                         30),
]

COLUMNS_REVIEW = [
    ("Référence",                   40),
    ("Année",                        8),
    ("Clé citable",                 15),
    ("BibTeX",                      30),
    ("Type",                        20),
    ("Objectif de la revue",        45),
    ("Corpus couvert",              35),
    ("Nb articles inclus",          15),
    ("Période couverte",            20),
    ("Thèmes principaux",           45),
    ("Consensus identifiés",        40),
    ("Débats / Controverses",       40),
    ("Limites identifiées",         35),
    ("Taille d'effet globale",      18),
    ("Hétérogénéité I²",           15),
    ("Take Home Message",           40),
    ("Domaines",                    20),
    ("Lu",                           8),
    ("PDF disponible",              15),
    ("Chemin PDF",                  40),
    ("Date ajout",                  15),
    ("PMID",                        12),
    ("DOI",                         30),
]


class ExcelManager:
    """Gestion du fichier Excel de veille scientifique."""

    def __init__(self, config: dict):
        self.excel_path = Path(config["paths"]["excel"])
        self.domains    = config["domains"]
        self.combos     = config.get("combinations", [])
        self.wb         = None

    # ── Initialisation / Chargement ──────────────────────────────────────────
    def load_or_create(self):
        """Charge le fichier Excel existant ou en crée un nouveau."""
        if self.excel_path.exists():
            self.wb = openpyxl.load_workbook(self.excel_path)
            logger.info(f"Excel chargé : {self.excel_path}")
            self._ensure_all_sheets()
        else:
            self.wb = openpyxl.Workbook()
            # Supprimer la feuille par défaut
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]
            self._create_all_sheets()
            self.save()
            logger.info(f"Nouveau fichier Excel créé : {self.excel_path}")

    def _sheet_names_expected(self) -> list[str]:
        """Retourne la liste de tous les onglets attendus."""
        names = ["Tous"]
        for d in self.domains:
            names.append(f"{d['short']} — Expérimental")
            names.append(f"{d['short']} — Revues")
        for c in self.combos:
            names.append(f"{c['short']} — Expérimental")
            names.append(f"{c['short']} — Revues")
        return names

    def _ensure_all_sheets(self):
        """Crée les onglets manquants dans un fichier existant."""
        for name in self._sheet_names_expected():
            if name not in self.wb.sheetnames:
                self._create_sheet(name)
                logger.info(f"Nouvel onglet créé : {name}")

    def _create_all_sheets(self):
        """Crée tous les onglets depuis zéro."""
        for name in self._sheet_names_expected():
            self._create_sheet(name)

    def _create_sheet(self, name: str):
        """Crée un onglet avec les colonnes appropriées."""
        ws = self.wb.create_sheet(title=name[:31])  # Excel: max 31 chars

        # Déterminer le type de colonnes
        is_review = "Revues" in name
        columns   = COLUMNS_REVIEW if is_review else COLUMNS_EXPERIMENTAL

        # Couleur de l'onglet
        if name == "Tous":
            ws.sheet_properties.tabColor = COLORS["tab_all"]
        elif "×" in name or "x" in name.lower() and "—" in name:
            ws.sheet_properties.tabColor = COLORS["tab_combined"]
        elif is_review:
            ws.sheet_properties.tabColor = COLORS["tab_review"]
        else:
            ws.sheet_properties.tabColor = COLORS["tab_experimental"]

        # En-tête
        self._write_header(ws, name, columns)

    def _write_header(self, ws: Worksheet, title: str, columns: list):
        """Écrit l'en-tête stylé sur un onglet."""
        # Ligne de titre (fusionnée)
        ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
        title_cell = ws["A1"]
        title_cell.value = f"🔬 Veille Scientifique — {title}"
        title_cell.font  = Font(
            bold=True, size=13, color=COLORS["header_fg"],
            name="Calibri"
        )
        title_cell.fill      = PatternFill("solid", fgColor=COLORS["header_bg"])
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Ligne d'en-tête des colonnes
        header_fill = PatternFill("solid", fgColor=COLORS["subheader_bg"])
        header_font = Font(bold=True, color=COLORS["header_fg"], name="Calibri", size=10)
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, (col_name, col_width) in enumerate(columns, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws.row_dimensions[2].height = 22

        # Figer les deux premières lignes
        ws.freeze_panes = "A3"

        # Filtre automatique
        ws.auto_filter.ref = f"A2:{get_column_letter(len(columns))}2"

    # ── Ajout d'un article ────────────────────────────────────────────────────
    def add_article(self, article: dict) -> bool:
        """
        Ajoute un article au fichier Excel dans les onglets appropriés.
        Vérifie les doublons par PMID ou DOI.
        """
        if self.wb is None:
            self.load_or_create()

        pmid = article.get("pmid", "")
        doi  = article.get("doi", "")

        # Vérification doublon global
        if self._is_duplicate(pmid, doi):
            logger.info(f"Article déjà présent (PMID: {pmid}) — ignoré")
            return False

        article_type  = article.get("article_type", "experimental")
        domains_codes = article.get("domains", [])

        # Onglets à remplir
        target_sheets = self._get_target_sheets(domains_codes, article_type)
        target_sheets.append("Tous")

        # Construction de la ligne
        row_data = self._build_row(article)

        for sheet_name in target_sheets:
            if sheet_name in self.wb.sheetnames:
                self._append_row(self.wb[sheet_name], row_data, article_type)

        self.save()
        logger.info(f"Article ajouté : {article.get('cite_key', 'inconnu')}")
        return True

    def _is_duplicate(self, pmid: str, doi: str) -> bool:
        """Vérifie si l'article est déjà dans l'onglet 'Tous'."""
        if "Tous" not in self.wb.sheetnames:
            return False
        ws = self.wb["Tous"]
        # Colonnes PMID et DOI (à la fin)
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not any(row):
                continue
            row_pmid = str(row[-2]) if len(row) >= 2 else ""
            row_doi  = str(row[-1]) if len(row) >= 1 else ""
            if pmid and row_pmid == pmid:
                return True
            if doi and row_doi == doi:
                return True
        return False

    def _get_target_sheets(self, domains_codes: list, article_type: str) -> list[str]:
        """Retourne la liste des onglets cibles pour un article."""
        suffix = "Revues" if article_type == "review" else "Expérimental"
        sheets = []

        # Onglets domaines simples
        for code in domains_codes:
            sheet_name = f"{code} — {suffix}"
            if sheet_name in self.wb.sheetnames:
                sheets.append(sheet_name)

        # Onglets combinaisons
        for combo in self.combos:
            combo_domains = set(combo["domains"])
            if combo_domains.issubset(set(domains_codes)):
                sheet_name = f"{combo['short']} — {suffix}"
                if sheet_name in self.wb.sheetnames:
                    sheets.append(sheet_name)

        return list(dict.fromkeys(sheets))  # Dédoublonnage ordre préservé

    def _safe_str(self, value) -> str:
        """Convertit n'importe quelle valeur en chaîne sûre pour Excel."""
        if value is None:
            return ""
        if isinstance(value, list):
            return " | ".join(str(v) for v in value)
        if isinstance(value, dict):
            return str(value)
        return str(value)

    def _build_row(self, article: dict) -> dict:
        """Construit le dict de données pour une ligne Excel."""
        analysis = article.get("analysis", {})
        art_type = article.get("article_type", "experimental")
        s = self._safe_str  # raccourci

        base = {
            "Référence":       s(article.get("reference", "")),
            "Année":           s(article.get("year", "")),
            "Clé citable":     s(article.get("cite_key", "")),
            "BibTeX":          s(article.get("bibtex", "")),
            "Type":            "Expérimental" if art_type == "experimental" else "Revue",
            "Domaines":        " | ".join(article.get("domains", [])),
            "Lu":              "Non",
            "PDF disponible":  "Oui" if article.get("pdf_available") else "Non",
            "Chemin PDF":      s(article.get("pdf_path", "")),
            "Date ajout":      datetime.now().strftime("%Y-%m-%d"),
            "PMID":            s(article.get("pmid", "")),
            "DOI":             s(article.get("doi", "")),
            "Take Home Message": s(analysis.get("take_home_message", "")),
        }

        if art_type == "experimental":
            base.update({
                "Hypothèse(s)":        s(analysis.get("hypotheses", "")),
                "Population":          s(analysis.get("population", "")),
                "N par groupe":        s(analysis.get("n_per_group", "")),
                "Type de groupe":      s(analysis.get("group_type", "")),
                "Critères incl./excl.":s(analysis.get("inclusion_criteria", "")),
                "Méthode / Outils":    s(analysis.get("methods", "")),
                "Résultats principaux":s(analysis.get("results", "")),
                "Taille d'effet":      s(analysis.get("effect_size", "")),
                "Conclusion":          s(analysis.get("conclusion", "")),
            })
        else:
            base.update({
                "Objectif de la revue":  s(analysis.get("review_objective", "")),
                "Corpus couvert":        s(analysis.get("corpus", "")),
                "Nb articles inclus":    s(analysis.get("n_articles", "")),
                "Période couverte":      s(analysis.get("period_covered", "")),
                "Thèmes principaux":     s(analysis.get("main_themes", "")),
                "Consensus identifiés":  s(analysis.get("consensus", "")),
                "Débats / Controverses": s(analysis.get("debates", "")),
                "Limites identifiées":   s(analysis.get("limitations", "")),
                "Taille d'effet globale":s(analysis.get("global_effect_size", "")),
                "Hétérogénéité I²":     s(analysis.get("heterogeneity", "")),
            })

        return base

    def _append_row(self, ws: Worksheet, row_data: dict, article_type: str):
        """Ajoute une ligne dans un onglet avec le style approprié."""
        columns = COLUMNS_REVIEW if article_type == "review" else COLUMNS_EXPERIMENTAL
        col_names = [c[0] for c in columns]

        row_idx = ws.max_row + 1
        is_even = (row_idx % 2 == 0)

        thin = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(col_names, 1):
            value = row_data.get(col_name, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font      = Font(name="Calibri", size=10)

            # Couleur de fond selon statut
            if col_name == "Lu":
                fill_color = COLORS["read"] if value == "Oui" else COLORS["unread"]
                cell.fill = PatternFill("solid", fgColor=fill_color)
            elif col_name == "PDF disponible":
                fill_color = COLORS["pdf_ok"] if value == "Oui" else COLORS["pdf_missing"]
                cell.fill = PatternFill("solid", fgColor=fill_color)
            else:
                fill_color = COLORS["row_even"] if is_even else COLORS["row_odd"]
                cell.fill = PatternFill("solid", fgColor=fill_color)

        ws.row_dimensions[row_idx].height = 60

    # ── Mise à jour d'un article (marquer comme lu, etc.) ────────────────────
    def mark_as_read(self, pmid: str):
        """Marque un article comme lu dans tous les onglets."""
        self._update_field(pmid, "Lu", "Oui", fill_color=COLORS["read"])
        self.save()

    def update_pdf_path(self, pmid: str, pdf_path: str):
        """Met à jour le chemin PDF d'un article."""
        self._update_field(pmid, "PDF disponible", "Oui", fill_color=COLORS["pdf_ok"])
        self._update_field(pmid, "Chemin PDF", pdf_path)
        self.save()

    def _update_field(self, pmid: str, field_name: str, value: str,
                      fill_color: Optional[str] = None):
        """Met à jour un champ dans tous les onglets pour un PMID donné."""
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            # Trouver la colonne du champ
            col_idx = None
            for cell in ws[2]:  # Ligne d'en-tête
                if cell.value == field_name:
                    col_idx = cell.column
                    break
            if col_idx is None:
                continue
            # Trouver la colonne PMID
            pmid_col = None
            for cell in ws[2]:
                if cell.value == "PMID":
                    pmid_col = cell.column
                    break
            if pmid_col is None:
                continue
            # Mettre à jour les lignes correspondantes
            for row in ws.iter_rows(min_row=3):
                if str(row[pmid_col - 1].value) == pmid:
                    target_cell = row[col_idx - 1]
                    target_cell.value = value
                    if fill_color:
                        target_cell.fill = PatternFill("solid", fgColor=fill_color)

    # ── Ajout d'un domaine à la volée ────────────────────────────────────────
    def add_domain(self, domain: dict, combos_to_add: list[dict] = None):
        """Ajoute un nouveau domaine et ses onglets."""
        self.domains.append(domain)
        self._create_sheet(f"{domain['short']} — Expérimental")
        self._create_sheet(f"{domain['short']} — Revues")
        if combos_to_add:
            for combo in combos_to_add:
                self.combos.append(combo)
                self._create_sheet(f"{combo['short']} — Expérimental")
                self._create_sheet(f"{combo['short']} — Revues")
        self.save()
        logger.info(f"Nouveau domaine ajouté : {domain['short']}")

    # ── Sauvegarde ───────────────────────────────────────────────────────────
    def save(self):
        """Sauvegarde le fichier Excel."""
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(self.excel_path)
        logger.debug(f"Excel sauvegardé : {self.excel_path}")

    # ── Statistiques ─────────────────────────────────────────────────────────
    def get_stats(self) -> dict:
        """Retourne des statistiques sur le fichier Excel."""
        if self.wb is None:
            return {}

        stats = {"total": 0, "by_domain": {}, "read": 0, "pdf_available": 0}
        if "Tous" not in self.wb.sheetnames:
            return stats

        ws = self.wb["Tous"]
        # Colonnes Lu et PDF disponible
        col_map = {}
        for cell in ws[2]:
            if cell.value in ("Lu", "PDF disponible", "Domaines"):
                col_map[cell.value] = cell.column

        for row in ws.iter_rows(min_row=3, values_only=True):
            if not any(row):
                continue
            stats["total"] += 1
            lu_val  = row[col_map.get("Lu", 1) - 1] if "Lu" in col_map else ""
            pdf_val = row[col_map.get("PDF disponible", 1) - 1] if "PDF disponible" in col_map else ""
            if lu_val == "Oui":
                stats["read"] += 1
            if pdf_val == "Oui":
                stats["pdf_available"] += 1

        return stats
