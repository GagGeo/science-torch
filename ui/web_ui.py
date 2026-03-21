"""
ui/web_ui.py — Interface web locale Science Torch
Serveur Flask léger qui s'ouvre dans le navigateur depuis le menu 🔬.
Pages : Tableau de bord + Configuration
"""
import sys as _sys
import os as _os
_sys.path.insert(0, _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))))

import json
import threading
import webbrowser
from pathlib import Path
from datetime import datetime

try:
    from flask import Flask, render_template_string, request, redirect
    HAS_FLASK = True
except ImportError:
    HAS_FLASK = False

from utils.logger import get_logger

logger = get_logger(__name__)
PORT = 7432

# ── Styles communs ────────────────────────────────────────────────────────────
STYLES = """
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');
:root {
  --bg:#0f1117; --surface:#1a1d27; --border:#2a2d3a;
  --accent:#4f9cf9; --accent2:#a78bfa; --success:#34d399;
  --warning:#fbbf24; --danger:#f87171; --text:#e2e8f0; --muted:#64748b;
  --mono:'DM Mono',monospace; --sans:'DM Sans',sans-serif;
}
*{box-sizing:border-box;margin:0;padding:0;}
body{background:var(--bg);color:var(--text);font-family:var(--sans);font-size:14px;line-height:1.6;min-height:100vh;}
.layout{display:flex;min-height:100vh;}
.sidebar{width:220px;background:var(--surface);border-right:1px solid var(--border);padding:24px 0;position:fixed;height:100vh;display:flex;flex-direction:column;}
.logo{padding:0 20px 24px;border-bottom:1px solid var(--border);margin-bottom:16px;}
.logo-icon{font-size:28px;display:block;margin-bottom:6px;}
.logo-name{font-size:13px;font-weight:600;letter-spacing:.08em;text-transform:uppercase;color:var(--accent);}
.logo-version{font-family:var(--mono);font-size:10px;color:var(--muted);margin-top:2px;}
.nav{flex:1;padding:0 12px;}
.nav-item{display:flex;align-items:center;gap:10px;padding:10px 12px;border-radius:8px;color:var(--muted);text-decoration:none;font-size:13px;font-weight:500;transition:all .15s;margin-bottom:2px;}
.nav-item:hover{background:var(--border);color:var(--text);}
.nav-item.active{background:rgba(79,156,249,.12);color:var(--accent);}
.nav-item .icon{font-size:16px;width:20px;text-align:center;}
.main{margin-left:220px;flex:1;padding:32px;max-width:1100px;}
.page-header{margin-bottom:32px;display:flex;align-items:flex-end;justify-content:space-between;}
.page-title{font-size:24px;font-weight:600;}
.page-subtitle{font-size:13px;color:var(--muted);margin-top:4px;}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;margin-bottom:32px;}
.card{background:var(--surface);border:1px solid var(--border);border-radius:12px;padding:20px;transition:border-color .2s;}
.card:hover{border-color:var(--accent);}
.card-label{font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.1em;color:var(--muted);margin-bottom:8px;}
.card-value{font-size:32px;font-weight:600;font-family:var(--mono);line-height:1;}
.card-sub{font-size:12px;color:var(--muted);margin-top:6px;}
.card-accent{border-color:var(--accent);}
.v-accent{color:var(--accent);} .v-success{color:var(--success);} .v-warning{color:var(--warning);}
.section{background:var(--surface);border:1px solid var(--border);border-radius:12px;margin-bottom:24px;overflow:hidden;}
.section-header{padding:16px 20px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between;}
.section-title{font-size:13px;font-weight:600;}
.section-body{padding:20px;}
.domain-row{display:flex;align-items:center;gap:12px;margin-bottom:12px;}
.domain-name{width:80px;font-family:var(--mono);font-size:12px;color:var(--accent);flex-shrink:0;}
.domain-bar-bg{flex:1;height:8px;background:var(--border);border-radius:4px;overflow:hidden;}
.domain-bar{height:100%;background:linear-gradient(90deg,var(--accent),var(--accent2));border-radius:4px;transition:width .8s cubic-bezier(.4,0,.2,1);}
.domain-count{font-family:var(--mono);font-size:12px;color:var(--muted);width:30px;text-align:right;}
.table{width:100%;border-collapse:collapse;}
.table th{font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.08em;color:var(--muted);padding:8px 12px;text-align:left;border-bottom:1px solid var(--border);}
.table td{padding:10px 12px;border-bottom:1px solid rgba(42,45,58,.5);font-size:13px;vertical-align:top;}
.table tr:last-child td{border-bottom:none;}
.table tr:hover td{background:rgba(255,255,255,.02);}
.badge{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:500;font-family:var(--mono);}
.b-blue{background:rgba(79,156,249,.15);color:var(--accent);}
.b-green{background:rgba(52,211,153,.15);color:var(--success);}
.b-red{background:rgba(248,113,113,.15);color:var(--danger);}
.b-purple{background:rgba(167,139,250,.15);color:var(--accent2);}
.form-group{margin-bottom:20px;}
.form-label{display:block;font-size:12px;font-weight:600;color:var(--muted);margin-bottom:6px;text-transform:uppercase;letter-spacing:.06em;}
.form-input{width:100%;background:var(--bg);border:1px solid var(--border);border-radius:8px;padding:10px 14px;color:var(--text);font-family:var(--sans);font-size:13px;transition:border-color .15s;outline:none;}
.form-input:focus{border-color:var(--accent);}
.form-row{display:grid;grid-template-columns:1fr 1fr;gap:16px;}
.domain-card{background:var(--bg);border:1px solid var(--border);border-radius:10px;padding:16px;margin-bottom:12px;}
.domain-card-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:8px;}
.domain-code{font-family:var(--mono);font-size:14px;font-weight:500;color:var(--accent);}
.keywords-list{display:flex;flex-wrap:wrap;gap:6px;}
.keyword-tag{background:rgba(79,156,249,.1);border:1px solid rgba(79,156,249,.2);color:var(--accent);padding:3px 10px;border-radius:20px;font-size:12px;font-family:var(--mono);}
.btn{display:inline-flex;align-items:center;gap:6px;padding:8px 16px;border-radius:8px;font-size:13px;font-weight:500;cursor:pointer;border:none;transition:all .15s;text-decoration:none;font-family:var(--sans);}
.btn-primary{background:var(--accent);color:#fff;}
.btn-primary:hover{background:#3b82f6;}
.btn-danger{background:rgba(248,113,113,.15);color:var(--danger);border:1px solid rgba(248,113,113,.3);}
.btn-danger:hover{background:rgba(248,113,113,.25);}
.btn-ghost{background:transparent;color:var(--muted);border:1px solid var(--border);}
.btn-ghost:hover{color:var(--text);}
.btn-sm{padding:5px 10px;font-size:12px;}
.tabs{display:flex;gap:4px;margin-bottom:24px;}
.tab{padding:8px 16px;border-radius:8px;font-size:13px;font-weight:500;cursor:pointer;border:1px solid transparent;color:var(--muted);background:transparent;transition:all .15s;font-family:var(--sans);}
.tab:hover{color:var(--text);}
.tab.active{background:var(--surface);border-color:var(--border);color:var(--text);}
.tab-panel{display:none;}
.tab-panel.active{display:block;}
.alert{padding:12px 16px;border-radius:8px;margin-bottom:16px;font-size:13px;}
.alert-success{background:rgba(52,211,153,.1);border:1px solid rgba(52,211,153,.2);color:var(--success);}
.alert-danger{background:rgba(248,113,113,.1);border:1px solid rgba(248,113,113,.2);color:var(--danger);}
.sidebar-footer{padding:16px 20px;border-top:1px solid var(--border);font-size:11px;color:var(--muted);font-family:var(--mono);}
.dot{width:8px;height:8px;border-radius:50%;display:inline-block;margin-right:6px;background:var(--success);box-shadow:0 0 6px var(--success);}
@keyframes fadeIn{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:translateY(0)}}
.main{animation:fadeIn .3s ease;}
"""

SIDEBAR = """
<aside class="sidebar">
  <div class="logo">
    <span class="logo-icon">🔬</span>
    <div class="logo-name">Science Torch</div>
    <div class="logo-version">v1.0.0</div>
  </div>
  <nav class="nav">
    <a href="/" class="nav-item {active_dash}"><span class="icon">📊</span> Dashboard</a>
    <a href="/config" class="nav-item {active_cfg}"><span class="icon">⚙️</span> Configuration</a>
  </nav>
  <div class="sidebar-footer"><span class="dot"></span>localhost:{port}</div>
</aside>
"""

JS = """
<script>
function switchTab(id){
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.querySelectorAll('.tab-panel').forEach(p=>p.classList.remove('active'));
  document.querySelector('[data-tab="'+id+'"]').classList.add('active');
  document.getElementById(id).classList.add('active');
}
document.addEventListener('DOMContentLoaded',()=>{
  document.querySelectorAll('.domain-bar').forEach(b=>{
    const w=b.dataset.width; b.style.width='0%';
    setTimeout(()=>b.style.width=w,100);
  });
});
</script>
"""

def page(title, sidebar, content):
    return f"""<!DOCTYPE html>
<html><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{title} — Science Torch</title>
<style>{STYLES}</style>
</head><body>
<div class="layout">
{sidebar}
<main class="main">{content}</main>
</div>
{JS}
</body></html>"""


class WebUI:
    def __init__(self, config: dict, config_path: Path):
        if not HAS_FLASK:
            raise ImportError("Flask non installé. Lancez : pip install flask")
        self.config      = config
        self.config_path = config_path
        self.app         = Flask(__name__)
        self.app.logger.disabled = True
        import logging
        log = logging.getLogger('werkzeug')
        log.setLevel(logging.ERROR)
        self._setup_routes()

    def _sidebar(self, active: str) -> str:
        return SIDEBAR.format(
            active_dash="active" if active == "dash" else "",
            active_cfg="active"  if active == "cfg"  else "",
            port=PORT
        )

    def _setup_routes(self):
        app = self.app

        # ── Dashboard ──────────────────────────────────────────────────────
        @app.route("/")
        def dashboard():
            stats, domain_stats, recent = self._get_dashboard_data()
            last = self._get_last_search()

            bars = "".join(
                f'<div class="domain-row">'
                f'<div class="domain-name">{d["code"]}</div>'
                f'<div class="domain-bar-bg"><div class="domain-bar" data-width="{d["pct"]}%" style="width:0%"></div></div>'
                f'<div class="domain-count">{d["count"]}</div>'
                f'</div>'
                for d in domain_stats
            ) or '<div style="color:var(--muted)">No data yet</div>'

            rows = "".join(
                f'<tr>'
                f'<td>{a["ref"]}</td>'
                f'<td><span class="badge b-blue">{a["year"]}</span></td>'
                '<td>' + ' '.join(f'<span class="badge b-purple">{d}</span>' for d in a['domains']) + '</td>'
                f'<td>{"<span class=\'badge b-green\'>✓ PDF</span>" if a["pdf"] else "<span class=\'badge b-red\'>✗</span>"}</td>'
                f'<td style="color:var(--muted);font-family:var(--mono);font-size:12px">{a["date"]}</td>'
                f'</tr>'
                for a in recent
            ) or '<tr><td colspan="5" style="color:var(--muted);text-align:center;padding:32px">No articles yet — run a search to get started</td></tr>'

            content = f"""
<div class="page-header">
  <div>
    <div class="page-title">Dashboard</div>
    <div class="page-subtitle">Last search: {last}</div>
  </div>
  <a href="/search" class="btn btn-primary">🔍 Run search now</a>
</div>
<div class="cards">
  <div class="card card-accent">
    <div class="card-label">Total articles</div>
    <div class="card-value v-accent">{stats['total']}</div>
    <div class="card-sub">in database</div>
  </div>
  <div class="card">
    <div class="card-label">PDFs available</div>
    <div class="card-value v-success">{stats['pdfs']}</div>
    <div class="card-sub">{stats['pdfs_pct']}% of total</div>
  </div>
  <div class="card">
    <div class="card-label">This week</div>
    <div class="card-value">{stats['this_week']}</div>
    <div class="card-sub">new articles</div>
  </div>
  <div class="card">
    <div class="card-label">Domains</div>
    <div class="card-value">{stats['domains']}</div>
    <div class="card-sub">monitored</div>
  </div>
</div>
<div class="section">
  <div class="section-header"><div class="section-title">📈 Articles by domain</div></div>
  <div class="section-body">{bars}</div>
</div>
<div class="section">
  <div class="section-header"><div class="section-title">🕐 Recent articles</div></div>
  <div class="section-body" style="padding:0">
    <table class="table">
      <thead><tr><th>Reference</th><th>Year</th><th>Domains</th><th>PDF</th><th>Added</th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
</div>"""
            return page("Dashboard", self._sidebar("dash"), content)

        # ── Config ────────────────────────────────────────────────────────
        @app.route("/config")
        def config_page():
            return self._render_config()

        @app.route("/config/domain/add", methods=["POST"])
        def add_domain():
            short    = request.form.get("short", "").strip().upper()
            name     = request.form.get("name", "").strip()
            keywords = [k.strip() for k in request.form.get("keywords", "").split(",") if k.strip()]
            if short and name and keywords:
                # Avoid duplicates
                codes = [d["short"] for d in self.config.get("domains", [])]
                if short not in codes:
                    self.config.setdefault("domains", []).append(
                        {"short": short, "name": name, "keywords": keywords}
                    )
                    self._save_config()
            return redirect("/config")

        @app.route("/config/domain/delete", methods=["POST"])
        def delete_domain():
            short = request.form.get("short", "").strip()
            self.config["domains"] = [d for d in self.config.get("domains", []) if d["short"] != short]
            self._save_config()
            return redirect("/config")

        @app.route("/config/settings", methods=["POST"])
        def save_settings():
            self.config.setdefault("ollama", {})["model"]    = request.form.get("ollama_model", "mistral")
            self.config["ollama"]["base_url"]                 = request.form.get("ollama_url", "http://localhost:11434")
            self.config.setdefault("scheduler", {})["day"]   = request.form.get("schedule_day", "monday")
            self.config["scheduler"]["time"]                  = request.form.get("schedule_time", "08:00")
            self.config["language"]                           = request.form.get("language", "en")
            self.config.setdefault("pubmed", {})["email"]     = request.form.get("pubmed_email", "")
            self.config["pubmed"]["max_results"]              = int(request.form.get("pubmed_max", 50) or 50)
            self._save_config()
            return redirect("/config")

        @app.route("/config/columns", methods=["POST"])
        def save_columns():
            try:
                exp_lines = [l.strip() for l in request.form.get("columns_exp", "").splitlines() if l.strip()]
                rev_lines = [l.strip() for l in request.form.get("columns_rev", "").splitlines() if l.strip()]
                self.config.setdefault("excel_columns", {})["experimental"] = [json.loads(l) for l in exp_lines]
                self.config["excel_columns"]["review"] = [json.loads(l) for l in rev_lines]
                self._save_config()
            except Exception as e:
                logger.error(f"Colonnes invalides : {e}")
            return redirect("/config")

        @app.route("/search")
        def trigger_search():
            def run():
                from core.scheduler import WeeklyScheduler
                WeeklyScheduler(self.config).run_weekly_search()
            threading.Thread(target=run, daemon=True).start()
            return redirect("/")

    def _render_config(self, msg=None, msg_type="success"):
        domains_html = ""
        for d in self.config.get("domains", []):
            kws = "".join(f'<span class="keyword-tag">{k}</span>' for k in d.get("keywords", []))
            domains_html += f"""
<div class="domain-card">
  <div class="domain-card-header">
    <span class="domain-code">{d['short']} — {d['name']}</span>
    <form method="POST" action="/config/domain/delete" style="display:inline">
      <input type="hidden" name="short" value="{d['short']}">
      <button type="submit" class="btn btn-danger btn-sm"
        onclick="return confirm('Delete {d['short']}?')">✕ Delete</button>
    </form>
  </div>
  <div class="keywords-list">{kws}</div>
</div>"""

        sched = self.config.get("scheduler", {})
        ollama = self.config.get("ollama", {})
        pubmed = self.config.get("pubmed", {})
        lang   = self.config.get("language", "en")

        days = ["monday","tuesday","wednesday","thursday","friday","saturday","sunday"]
        day_opts = "".join(
            f'<option value="{d}" {"selected" if sched.get("day") == d else ""}>{d.capitalize()}</option>'
            for d in days
        )
        lang_opts = f"""
<option value="en" {"selected" if lang == "en" else ""}>English</option>
<option value="fr" {"selected" if lang == "fr" else ""}>Français</option>"""

        exp_cols = self.config.get("excel_columns", {}).get("experimental", [])
        rev_cols = self.config.get("excel_columns", {}).get("review", [])
        exp_json = "\n".join(json.dumps(c, ensure_ascii=False) for c in exp_cols)
        rev_json = "\n".join(json.dumps(c, ensure_ascii=False) for c in rev_cols)

        alert_html = f'<div class="alert alert-{msg_type}">{msg}</div>' if msg else ""

        content = f"""
<div class="page-header">
  <div>
    <div class="page-title">Configuration</div>
    <div class="page-subtitle">Customize Science Torch for your research</div>
  </div>
</div>
{alert_html}
<div class="tabs">
  <button class="tab active" data-tab="tab-domains" onclick="switchTab('tab-domains')">Domains</button>
  <button class="tab" data-tab="tab-settings" onclick="switchTab('tab-settings')">Settings</button>
  <button class="tab" data-tab="tab-columns" onclick="switchTab('tab-columns')">Excel columns</button>
</div>

<div id="tab-domains" class="tab-panel active">
{domains_html}
<div class="section" style="margin-top:24px">
  <div class="section-header"><div class="section-title">➕ Add a domain</div></div>
  <div class="section-body">
    <form method="POST" action="/config/domain/add">
      <div class="form-row">
        <div class="form-group">
          <label class="form-label">Short code (e.g. ME)</label>
          <input type="text" name="short" class="form-input" placeholder="ME" required>
        </div>
        <div class="form-group">
          <label class="form-label">Full name</label>
          <input type="text" name="name" class="form-input" placeholder="Episodic Memory" required>
        </div>
      </div>
      <div class="form-group">
        <label class="form-label">PubMed keywords (comma-separated)</label>
        <input type="text" name="keywords" class="form-input" placeholder="episodic memory, memory encoding" required>
      </div>
      <button type="submit" class="btn btn-primary">Add domain</button>
    </form>
  </div>
</div>
</div>

<div id="tab-settings" class="tab-panel">
<form method="POST" action="/config/settings">
  <div class="section">
    <div class="section-header"><div class="section-title">🤖 Ollama</div></div>
    <div class="section-body">
      <div class="form-row">
        <div class="form-group">
          <label class="form-label">Model</label>
          <input type="text" name="ollama_model" class="form-input" value="{ollama.get('model','mistral')}">
        </div>
        <div class="form-group">
          <label class="form-label">Base URL</label>
          <input type="text" name="ollama_url" class="form-input" value="{ollama.get('base_url','http://localhost:11434')}">
        </div>
      </div>
    </div>
  </div>
  <div class="section">
    <div class="section-header"><div class="section-title">🗓 Schedule</div></div>
    <div class="section-body">
      <div class="form-row">
        <div class="form-group">
          <label class="form-label">Day</label>
          <select name="schedule_day" class="form-input">{day_opts}</select>
        </div>
        <div class="form-group">
          <label class="form-label">Time (HH:MM)</label>
          <input type="text" name="schedule_time" class="form-input" value="{sched.get('time','08:00')}">
        </div>
      </div>
    </div>
  </div>
  <div class="section">
    <div class="section-header"><div class="section-title">🌐 Language</div></div>
    <div class="section-body">
      <div class="form-group">
        <label class="form-label">Interface language</label>
        <select name="language" class="form-input">{lang_opts}</select>
      </div>
    </div>
  </div>
  <div class="section">
    <div class="section-header"><div class="section-title">📚 PubMed</div></div>
    <div class="section-body">
      <div class="form-row">
        <div class="form-group">
          <label class="form-label">Email (NCBI, optional)</label>
          <input type="email" name="pubmed_email" class="form-input" value="{pubmed.get('email','')}">
        </div>
        <div class="form-group">
          <label class="form-label">Max results per search</label>
          <input type="number" name="pubmed_max" class="form-input" value="{pubmed.get('max_results',50)}">
        </div>
      </div>
    </div>
  </div>
  <button type="submit" class="btn btn-primary">💾 Save settings</button>
</form>
</div>

<div id="tab-columns" class="tab-panel">
<form method="POST" action="/config/columns">
  <div class="section">
    <div class="section-header"><div class="section-title">🧪 Experimental / Meta-analysis</div></div>
    <div class="section-body">
      <textarea name="columns_exp" class="form-input" rows="12"
        style="font-family:var(--mono);font-size:12px;resize:vertical">{exp_json}</textarea>
      <div style="color:var(--muted);font-size:12px;margin-top:8px">
        One column per line — {{"name": "Column name", "width": 30}}
      </div>
    </div>
  </div>
  <div class="section">
    <div class="section-header"><div class="section-title">📖 Literature reviews</div></div>
    <div class="section-body">
      <textarea name="columns_rev" class="form-input" rows="12"
        style="font-family:var(--mono);font-size:12px;resize:vertical">{rev_json}</textarea>
    </div>
  </div>
  <button type="submit" class="btn btn-primary">💾 Save columns</button>
</form>
</div>"""

        return page("Configuration", self._sidebar("cfg"), content)

    def _get_dashboard_data(self):
        excel_path = Path(self.config.get("paths", {}).get("excel", ""))
        stats = {"total":0,"pdfs":0,"pdfs_pct":0,"this_week":0,"domains":len(self.config.get("domains",[]))}
        domain_stats, recent = [], []

        if not excel_path.exists():
            return stats, domain_stats, recent

        try:
            import openpyxl
            from datetime import date, timedelta
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet_name = next((s for s in ["Tous","All"] if s in wb.sheetnames), None)
            if not sheet_name:
                wb.close(); return stats, domain_stats, recent

            ws = wb[sheet_name]
            headers = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]

            def ci(name, alt=None):
                try: return headers.index(name)
                except ValueError:
                    if alt:
                        try: return headers.index(alt)
                        except ValueError: pass
                    return None

            ref_i  = ci("Référence","Reference") or 0
            dom_i  = ci("Domaines","Domains")
            pdf_i  = ci("PDF disponible","PDF available")
            date_i = ci("Date ajout","Date added")
            year_i = ci("Année","Year") or 1
            week_ago = (date.today() - timedelta(days=7)).strftime("%Y-%m-%d")
            domain_counts = {}

            for row in ws.iter_rows(min_row=3, values_only=True):
                if not any(row): continue
                stats["total"] += 1
                if pdf_i is not None and row[pdf_i] in ("Oui","Yes",True,"TRUE","1"): stats["pdfs"] += 1
                if date_i is not None and row[date_i] and str(row[date_i]) >= week_ago: stats["this_week"] += 1
                if dom_i is not None and row[dom_i]:
                    for d in str(row[dom_i]).split("|"):
                        d = d.strip()
                        if d: domain_counts[d] = domain_counts.get(d,0)+1
                if len(recent) < 10:
                    doms = [d.strip() for d in str(row[dom_i] or "").split("|") if d.strip()] if dom_i is not None else []
                    recent.append({
                        "ref":     str(row[ref_i] or "")[:60],
                        "year":    str(row[year_i] or ""),
                        "domains": doms,
                        "pdf":     pdf_i is not None and row[pdf_i] in ("Oui","Yes",True),
                        "date":    str(row[date_i] or "")[:10] if date_i is not None else "",
                    })

            if stats["total"] > 0:
                stats["pdfs_pct"] = round(stats["pdfs"]/stats["total"]*100)
                mx = max(domain_counts.values()) if domain_counts else 1
                domain_stats = [
                    {"code":k,"count":v,"pct":round(v/mx*100)}
                    for k,v in sorted(domain_counts.items(),key=lambda x:-x[1])
                ]
            wb.close()
        except Exception as e:
            logger.error(f"Dashboard error: {e}")

        return stats, domain_stats, recent

    def _get_last_search(self):
        p = Path(self.config.get("paths",{}).get("summaries",""))
        if p.exists():
            files = sorted(p.glob("summary_*.md"),reverse=True) or sorted(p.glob("resume_*.md"),reverse=True)
            if files:
                ts = files[0].stem.replace("summary_","").replace("resume_","")
                try: return datetime.strptime(ts,"%Y_%m_%d").strftime("%B %d, %Y")
                except: pass
        return "Never"

    def _save_config(self):
        with open(self.config_path,"w",encoding="utf-8") as f:
            json.dump(self.config,f,ensure_ascii=False,indent=2)

    def open_browser(self):
        def _open():
            import time; time.sleep(1.2)
            webbrowser.open(f"http://localhost:{PORT}")
        threading.Thread(target=_open,daemon=True).start()

    def run(self, open_browser=True):
        if open_browser: self.open_browser()
        logger.info(f"Interface web démarrée sur http://localhost:{PORT}")
        self.app.run(host="127.0.0.1",port=PORT,debug=False,use_reloader=False)
